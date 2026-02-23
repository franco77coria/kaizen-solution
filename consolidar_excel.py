"""
Script de consolidación de archivos Excel → CSV
Consolida todos los .xlsx de "carpetas a trabajar" en un único CSV
con la estructura del Dashboard (14 columnas).
"""

import openpyxl
import csv
import os
import re
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning)

# ─── CONFIGURACIÓN ───────────────────────────────────────────────────────────
BASE_DIR = r'c:\Users\Usuario\OneDrive\Escritorio\Pagina web\carpetas a trabajar'
OUTPUT_CSV = r'c:\Users\Usuario\OneDrive\Escritorio\Pagina web\consolidado_dashboard.csv'

CSV_COLUMNS = [
    'NOMBRE_LIDER', 'CC_LIDER', 'RELACION_PARENTESCO', 'A_QUIEN_REFIERE',
    'NUMERO_DOCUMENTO', 'OCUPACION_NEGOCIO', 'TELEFONO', 'DEPARTAMENTO',
    'MUNICIPIO', 'FECHA_NACIMIENTO', 'USA_WHATSAPP', 'FECHA_AUTORIZACION',
    'SEXO', 'OBSERVACIONES'
]

# Sheets to skip (lookup tables, pivot tables, empty)
SKIP_SHEETS = ['hoja2', 'tabla dinámica 1', 'tabla dinßmica 1', 'hoja1']

# Stats
stats = {
    'files_processed': 0,
    'files_skipped': 0,
    'sheets_processed': 0,
    'sheets_skipped': 0,
    'rows_added': 0,
    'rows_empty_skipped': 0,
    'errors': [],
}


# ─── UTILITIES ────────────────────────────────────────────────────────────────

def clean_val(v):
    """Clean a cell value to a string, stripping whitespace."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ('', 'None', 'none', '0', '0.0'):
        return None
    # Remove trailing .0 from numbers read as float
    if s.endswith('.0') and s.replace('.', '').replace('-', '').isdigit():
        s = s[:-2]
    return s


def clean_number(v):
    """Clean a numeric value (cedula, telefono)."""
    if v is None:
        return None
    s = str(v).strip().replace(',', '').replace(' ', '')
    if s in ('', 'None', 'none', '0', '0.0', '#VALUE!', '#REF!'):
        return None
    # Remove trailing .0
    if '.' in s:
        try:
            num = float(s)
            if num == int(num):
                s = str(int(num))
            else:
                s = str(num)
        except ValueError:
            pass
    # Remove leading 57 country code from phone-like numbers in CON INDI column
    return s


def clean_phone(v):
    """Clean a phone number, removing country code prefix."""
    s = clean_number(v)
    if s is None:
        return None
    # Remove 57 prefix if present (Colombian country code)
    if len(s) > 10 and s.startswith('57'):
        s = s[2:]
    return s


def build_full_name(*parts):
    """Build a full name from parts, handling None values."""
    valid = [str(p).strip() for p in parts if p is not None and str(p).strip() not in ('', 'None', '0', '0.0')]
    return ' '.join(valid).upper() if valid else None


def make_row(nombre_lider=None, cc_lider=None, relacion=None, a_quien=None,
             num_doc=None, ocupacion='Persona Natural', telefono=None,
             departamento='Cundinamarca', municipio=None, fecha_nac=None,
             usa_wa='Sí', fecha_aut=None, sexo=None, observaciones=None):
    """Create a standardized row dict."""
    return {
        'NOMBRE_LIDER': nombre_lider or 'null',
        'CC_LIDER': cc_lider or 'null',
        'RELACION_PARENTESCO': relacion or 'null',
        'A_QUIEN_REFIERE': a_quien or 'null',
        'NUMERO_DOCUMENTO': num_doc or 'null',
        'OCUPACION_NEGOCIO': ocupacion,
        'TELEFONO': telefono or 'null',
        'DEPARTAMENTO': departamento,
        'MUNICIPIO': municipio or 'null',
        'FECHA_NACIMIENTO': fecha_nac or 'null',
        'USA_WHATSAPP': usa_wa,
        'FECHA_AUTORIZACION': fecha_aut or 'null',
        'SEXO': sexo or 'null',
        'OBSERVACIONES': observaciones or 'SE ENVIO MENSAJE',
    }


def extract_lider_from_filename(filename):
    """
    Extract NOMBRE_LIDER from filename.
    Examples:
      'REFERIDOS YOLANDA URBINA.xlsx' → 'YOLANDA URBINA'
      'Lista Ingrid Gonzalez.xlsx' → 'INGRID GONZALEZ'
      'planillas alex prieto (RAFAEL VILLATE).xlsx' → 'RAFAEL VILLATE'
      'Copia de planilla A.P cota NICOLÁS CALDERON.xlsx' → 'NICOLÁS CALDERON'
      'Cristian muñoz- planilla A.P cota.xlsx' → 'CRISTIAN MUÑOZ'
      'PLANILLA BASE DATOS ANDREA RODRIGUEZ RUBEN CACERES.xlsx' → 'ANDREA RODRIGUEZ RUBEN CACERES'
      'Johanna Bermúdez Bomberos.xlsx' → 'JOHANNA BERMÚDEZ'
      'REFERENTES ANA JENIFFER GONZALEZ RODRIGGUEZ.xlsx' → 'ANA JENIFFER GONZALEZ RODRIGGUEZ'
      'Girardot.xlsx' → None (just a municipality)
    """
    name = os.path.splitext(filename)[0]

    # If has parentheses, extract content inside them
    paren = re.search(r'\(([^)]+)\)', name)
    if paren:
        return paren.group(1).strip().upper()

    # Pattern: "REFERIDOS <NAME>" or "REFERENTES <NAME>"
    m = re.match(r'(?:REFERIDOS?|REFERENTES?)\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "Lista <NAME>"
    m = re.match(r'Lista\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "PLANILLA BASE DATOS <NAME>"
    m = re.match(r'PLANILLA\s+BASE\s+DATOS\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "<Name>- planilla ..." or "<Name> planilla..."
    m = re.match(r'([A-Za-zÀ-ÿ\s]+?)[\s-]+planilla', name, re.IGNORECASE)
    if m:
        candidate = m.group(1).strip()
        if candidate.lower() not in ('copia de', 'copia'):
            return candidate.upper()

    # Pattern: "Copia de planilla A.P cota <NAME>"
    m = re.match(r'Copia\s+de\s+planilla.*?cota\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "LISTADO CAMPAÑA CAMARA ALEX <NAME>"
    m = re.match(r'LISTADO\s+CAMPAÑA\s+CAMARA\s+ALEX\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    # Same with encoding issues
    m = re.match(r'LISTADO\s+CAMPA.*?\s+CAMARA\s+ALEX\s+(.+)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "<Name> Bomberos" or "<Name> <descriptor>"
    m = re.match(r'^([A-Za-zÀ-ÿ]+\s+[A-Za-zÀ-ÿ]+)\s+(?:Bomberos)', name, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()

    # Pattern: "Copia de planilla <NAME>"
    m = re.match(r'Copia\s+de\s+planilla\s+(.+)', name, re.IGNORECASE)
    if m:
        candidate = m.group(1).strip()
        # Remove "A.P cota" or similar
        candidate = re.sub(r'A\.?P\.?\s*cota\s*', '', candidate, flags=re.IGNORECASE).strip()
        if candidate:
            return candidate.upper()

    return None


def extract_municipio_from_filename(filename):
    """Extract municipality from filename for simple municipality files."""
    name = os.path.splitext(filename)[0]
    # Remove trailing underscores
    name = name.rstrip('_').strip()
    return name.upper()


def get_cell(ws, row, col):
    """Get cell value from worksheet."""
    return ws.cell(row=row, column=col).value


def find_header_row(ws, max_search=15):
    """Find the row containing headers by looking for known header keywords."""
    keywords = {'nombre', 'apellido', 'cedula', 'celular', 'telefono', 'número',
                'cédula', 'n°', 'no.', 'nombre completo', 'nombres', 'apellidos'}
    for r in range(1, min(max_search, ws.max_row + 1)):
        row_text = set()
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                row_text.add(str(v).strip().lower().rstrip(' '))
        matches = row_text.intersection(keywords)
        if len(matches) >= 2:
            return r
    return None


def detect_and_parse_sheet(ws, sheet_name, filename, filepath):
    """
    Detect the type of sheet and parse all data rows.
    Returns a list of row dicts.
    """
    rows = []
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2:
        return rows

    # Read first 15 rows to detect structure
    sample = {}
    for r in range(1, min(16, max_row + 1)):
        sample[r] = {}
        for c in range(1, min(max_col + 1, 30)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                sample[r][c] = str(v).strip()

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE A: Province files (RESPONSABLE column)
    # Headers like: control, CELULAR, RESPONSABLE, No., APELLIDOS, NOMBRES, CEDULA, SEXO, ...
    # or: control, RESPONSABLE, No., APELLIDOS, NOMBRES, NOMBRE, CEDULA, CELULAR, ...
    # ═══════════════════════════════════════════════════════════════════════
    header_row_1 = sample.get(1, {})
    header_text_1 = ' '.join(header_row_1.values()).lower()

    if 'responsable' in header_text_1:
        # Find column positions
        col_map = {}
        for c, v in header_row_1.items():
            vl = v.lower().strip()
            if 'responsable' in vl:
                col_map['responsable'] = c
            elif vl in ('apellidos', 'apellidos '):
                col_map['apellidos'] = c
            elif vl in ('nombres', 'nombres '):
                col_map['nombres'] = c
            elif vl in ('nombre',):
                col_map['nombre_completo'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c
            elif 'celular' in vl or 'telefono' in vl or vl == 'numero':
                col_map['celular'] = c
            elif vl in ('sexo',):
                col_map['sexo'] = c
            elif 'municipio' in vl:
                col_map['municipio'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c

        for r in range(2, max_row + 1):
            responsable = clean_val(get_cell(ws, r, col_map.get('responsable', 99)))
            apellidos = clean_val(get_cell(ws, r, col_map.get('apellidos', 99)))
            nombres = clean_val(get_cell(ws, r, col_map.get('nombres', 99)))
            nombre_completo = clean_val(get_cell(ws, r, col_map.get('nombre_completo', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            sexo_raw = clean_val(get_cell(ws, r, col_map.get('sexo', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))

            # Build full name
            if nombre_completo:
                a_quien = nombre_completo.upper()
            else:
                a_quien = build_full_name(nombres, apellidos)

            if not a_quien and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            # Map sexo
            sexo = None
            if sexo_raw:
                sl = sexo_raw.lower()
                if sl in ('f', 'femenino'):
                    sexo = 'Femenino'
                elif sl in ('m', 'masculino'):
                    sexo = 'Masculino'

            # Clean municipio (remove 0.0 values)
            if municipio and municipio in ('0', '0.0'):
                municipio = None

            rows.append(make_row(
                nombre_lider=responsable.upper() if responsable else None,
                a_quien=a_quien,
                num_doc=cedula,
                telefono=celular,
                municipio=municipio,
                sexo=sexo,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE B: Planilla files (LIDER at R7, headers at R8-R9, data at R11+)
    # ═══════════════════════════════════════════════════════════════════════
    r7_text = ' '.join(sample.get(7, {}).values()).lower() if 7 in sample else ''
    r8_text = ' '.join(sample.get(8, {}).values()).lower() if 8 in sample else ''
    r9_text = ' '.join(sample.get(9, {}).values()).lower() if 9 in sample else ''

    if 'lider' in r7_text or ('nombre completo' in r9_text) or ('nombre completo' in r8_text):
        # Extract LIDER name from R7
        lider = None
        r7_vals = sample.get(7, {})
        for c, v in r7_vals.items():
            if 'lider' in v.lower():
                # Extract name after "LIDER :" or "LIDER:"
                m = re.search(r'LIDER\s*:\s*(.+)', v, re.IGNORECASE)
                if m:
                    lider_name = m.group(1).strip()
                    if lider_name:
                        lider = lider_name.upper()
                break

        # If no lider from R7, try filename
        if not lider:
            lider = extract_lider_from_filename(filename)

        # Determine header row (R8 or R9) and data start
        if 'nombre completo' in r9_text:
            header_r = 9
            data_start = 11  # R10 is sub-header (DD/MM/AA), data at R11
        elif 'nombre completo' in r8_text:
            header_r = 8
            data_start = 10
        else:
            # Try to find it
            header_r = find_header_row(ws, 15)
            if header_r:
                data_start = header_r + 2
            else:
                return rows

        # Map columns from header row
        col_map = {}
        for c in range(1, min(max_col + 1, 20)):
            v = ws.cell(row=header_r, column=c).value
            if v is None:
                continue
            vl = str(v).strip().lower()
            if 'nombre completo' in vl:
                col_map['nombre'] = c
            elif vl in ('n° cedula', 'n° cedula', 'nro cedula', 'nº cedula', 'cedula'):
                col_map['cedula'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c
            elif 'celular' in vl or 'telefono' in vl:
                col_map['celular'] = c
            elif 'parentesco' in vl:
                col_map['parentesco'] = c
            elif 'municipio' in vl:
                col_map['municipio'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'votos' in vl:
                col_map['votos'] = c  # skip this column
            elif 'llamada' in vl:
                col_map['observaciones'] = c

        for r in range(data_start, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            parentesco = clean_val(get_cell(ws, r, col_map.get('parentesco', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))

            if not nombre and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=lider,
                a_quien=nombre.upper() if nombre else None,
                num_doc=cedula,
                telefono=celular,
                relacion=parentesco,
                municipio=municipio,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE C: Consolidado Cota files
    # Headers: No./number, Numero(phone), MUNICIPIO RESIDENCIA, APELLIDOS, NOMBRES, CON INDI, OBSERVACIONES
    # ═══════════════════════════════════════════════════════════════════════
    if 'municipio residencia' in header_text_1:
        col_map = {}
        for c, v in header_row_1.items():
            vl = v.lower().strip()
            if 'municipio' in vl:
                col_map['municipio'] = c
            elif vl in ('apellidos', 'apellidos '):
                col_map['apellidos'] = c
            elif vl in ('nombres', 'nombres '):
                col_map['nombres'] = c
            elif vl in ('numero', 'número'):
                col_map['celular'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c

        file_lider = extract_lider_from_filename(filename)

        for r in range(2, max_row + 1):
            apellidos = clean_val(get_cell(ws, r, col_map.get('apellidos', 99)))
            nombres = clean_val(get_cell(ws, r, col_map.get('nombres', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))

            a_quien = build_full_name(nombres, apellidos)

            if not a_quien and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=file_lider,
                a_quien=a_quien,
                num_doc=cedula,
                telefono=celular,
                municipio=municipio,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE D: REVISAR-BASE DE DATOS REFERIDOS DEPURADA (BD sheet)
    # Headers: NOMBRES Y APELLIDOS, cedula, celular, Ciudad, RESPONSABLE, ...
    # ═══════════════════════════════════════════════════════════════════════
    if 'nombres y apellidos' in header_text_1 or 'nombre completo' in header_text_1:
        col_map = {}
        for c, v in header_row_1.items():
            vl = v.lower().strip()
            if 'nombre' in vl:
                col_map['nombre'] = c
            elif 'responsable' in vl:
                col_map['responsable'] = c
            elif 'celular' in vl:
                col_map['celular'] = c
            elif 'ciudad' in vl or 'municipio' in vl:
                col_map['municipio'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'cédula' in vl or 'cedula' in vl:
                col_map['cedula'] = c
            elif 'contesto' in vl or 'llamada' in vl:
                col_map['llamada'] = c
            elif 'vereda' in vl:
                col_map['vereda'] = c  # skip

        # Check if first item in R1 looks like data (already has a cedula-like value in C2)
        # In some files R1 IS also data (the header is implicit)
        r1_c2 = header_row_1.get(2, '')
        has_explicit_header = not r1_c2.replace(',', '').replace('.', '').isdigit()

        start_row = 2 if has_explicit_header else 1
        file_lider = extract_lider_from_filename(filename)

        for r in range(start_row, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            responsable = clean_val(get_cell(ws, r, col_map.get('responsable', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            llamada = clean_val(get_cell(ws, r, col_map.get('llamada', 99)))

            if not nombre and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            lider = responsable.upper() if responsable else file_lider
            final_obs = obs or llamada or 'SE ENVIO MENSAJE'

            rows.append(make_row(
                nombre_lider=lider,
                a_quien=nombre.upper() if nombre else None,
                num_doc=cedula,
                telefono=celular,
                municipio=municipio,
                observaciones=final_obs,
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE E: REVISAR-BASE CONSOLIDADO sheet (headers at R7)
    # ═══════════════════════════════════════════════════════════════════════
    r7_vals = sample.get(7, {})
    r7_text_all = ' '.join(r7_vals.values()).lower() if r7_vals else ''
    if 'nombre completo' in r7_text_all and 'refrenciado' in r7_text_all:
        col_map = {}
        for c, v in r7_vals.items():
            vl = v.lower().strip()
            if 'nombre completo' in vl:
                col_map['nombre'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c
            elif 'celular' in vl:
                col_map['celular'] = c
            elif 'parentesco' in vl:
                col_map['parentesco'] = c
            elif 'municipio' in vl:
                col_map['municipio'] = c
            elif 'refrenciado' in vl or 'referenciado' in vl:
                col_map['responsable'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'contesto' in vl:
                col_map['llamada'] = c

        for r in range(8, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            parentesco = clean_val(get_cell(ws, r, col_map.get('parentesco', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            responsable = clean_val(get_cell(ws, r, col_map.get('responsable', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
            llamada = clean_val(get_cell(ws, r, col_map.get('llamada', 99)))

            if not nombre and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=responsable.upper() if responsable else None,
                a_quien=nombre.upper() if nombre else None,
                num_doc=cedula,
                telefono=celular,
                relacion=parentesco,
                municipio=municipio,
                observaciones=obs or llamada or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE F: SIN DUPLICIDADES sheet (headers at R7)
    # TELEFONO, Nombre, Tema de interes, REFRENCIADO
    # ═══════════════════════════════════════════════════════════════════════
    if 'telefono' in r7_text_all and 'refrenciado' in r7_text_all:
        col_map = {}
        for c, v in r7_vals.items():
            vl = v.lower().strip()
            if 'nombre' in vl:
                col_map['nombre'] = c
            elif 'telefono' in vl:
                col_map['celular'] = c
            elif 'refrenciado' in vl:
                col_map['responsable'] = c

        for r in range(8, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            responsable = clean_val(get_cell(ws, r, col_map.get('responsable', 99)))

            if not nombre and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=responsable.upper() if responsable else None,
                a_quien=nombre.upper() if nombre else None,
                telefono=celular,
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE G: Event files / special format with headers in R2
    # (REUNION, BRIGADA) - headers at row 2, data from row 3
    # ═══════════════════════════════════════════════════════════════════════
    r2_vals = sample.get(2, {})
    r2_text = ' '.join(r2_vals.values()).lower() if r2_vals else ''
    if ('nombre' in r2_text and ('apellido' in r2_text or 'celular' in r2_text)):
        col_map = {}
        for c, v in r2_vals.items():
            vl = v.lower().strip()
            if vl in ('nombre', 'nombre '):
                col_map['nombre'] = c
            elif 'apellido' in vl:
                col_map['apellido'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c
            elif 'celular' in vl:
                col_map['celular'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'vereda' in vl:
                col_map['vereda'] = c  # can use as municipio hint

        file_lider = extract_lider_from_filename(filename)
        # Try to get municipio from title in R1
        r1_title = ' '.join(sample.get(1, {}).values())
        event_municipio = None
        if 'cota' in r1_title.lower():
            event_municipio = 'COTA'

        for r in range(3, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            apellido = clean_val(get_cell(ws, r, col_map.get('apellido', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))

            a_quien = build_full_name(nombre, apellido)

            if not a_quien and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=file_lider,
                a_quien=a_quien,
                num_doc=cedula,
                telefono=celular,
                municipio=event_municipio,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE H: Bomberos/Special - headers at R3
    # CARGO, NOMBRE, APELLIDO, CEDULA, TELEFONO, CORREO, FECHA NACIMIENTO, OBSERVACIONES
    # ═══════════════════════════════════════════════════════════════════════
    r3_vals = sample.get(3, {})
    r3_text = ' '.join(r3_vals.values()).lower() if r3_vals else ''
    if 'cargo' in r3_text and 'nombre' in r3_text:
        col_map = {}
        for c, v in r3_vals.items():
            vl = v.lower().strip()
            if vl in ('nombre', 'nombre '):
                col_map['nombre'] = c
            elif 'apellido' in vl:
                col_map['apellido'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c
            elif 'telefono' in vl or 'celular' in vl:
                col_map['celular'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'fecha' in vl and 'nacimiento' in vl:
                col_map['fecha_nac'] = c

        file_lider = extract_lider_from_filename(filename)

        for r in range(4, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            apellido = clean_val(get_cell(ws, r, col_map.get('apellido', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
            fecha_nac_raw = get_cell(ws, r, col_map.get('fecha_nac', 99))

            a_quien = build_full_name(nombre, apellido)
            if not a_quien and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            fecha_nac = None
            if fecha_nac_raw:
                try:
                    if hasattr(fecha_nac_raw, 'strftime'):
                        fecha_nac = fecha_nac_raw.strftime('%Y-%m-%d')
                    else:
                        fecha_nac = str(fecha_nac_raw)
                except:
                    pass

            rows.append(make_row(
                nombre_lider=file_lider,
                a_quien=a_quien,
                num_doc=cedula,
                telefono=celular,
                fecha_nac=fecha_nac,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE I: CAMINATAS / NUMERO, NOMBRE, APELLIDO, TELEFONO, MUNICIPIO
    # Headers at R1 with columns: NUMERO, NOMBRE, APELLIDO, TELEFONO, MUNICIPIO, OBSERVACIONES
    # ═══════════════════════════════════════════════════════════════════════
    if 'numero' in header_text_1 and 'nombre' in header_text_1:
        col_map = {}
        for c, v in header_row_1.items():
            vl = v.lower().strip()
            if vl in ('nombre', 'nombre '):
                col_map['nombre'] = c
            elif 'apellido' in vl:
                col_map['apellido'] = c
            elif 'telefono' in vl or 'celular' in vl:
                col_map['celular'] = c
            elif 'municipio' in vl:
                col_map['municipio'] = c
            elif 'observaciones' in vl:
                col_map['observaciones'] = c
            elif 'cedula' in vl or 'cédula' in vl:
                col_map['cedula'] = c

        file_lider = extract_lider_from_filename(filename)

        for r in range(2, max_row + 1):
            nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
            apellido = clean_val(get_cell(ws, r, col_map.get('apellido', 99)))
            celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
            municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
            obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
            cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))

            a_quien = build_full_name(nombre, apellido)

            if not a_quien and not celular:
                stats['rows_empty_skipped'] += 1
                continue

            rows.append(make_row(
                nombre_lider=file_lider,
                a_quien=a_quien,
                num_doc=cedula,
                telefono=celular,
                municipio=municipio,
                observaciones=obs or 'SE ENVIO MENSAJE',
            ))
        return rows

    # ═══════════════════════════════════════════════════════════════════════
    # TYPE J: Simple municipality files (Nombre, Apellido, Cedula, Celular, Llamada, Observaciones)
    # Also: REFERENTES with NOMBRE, APELLIDO, CÉDULA, TELEFONO, MUNICIPIO
    # Auto-detect header row
    # ═══════════════════════════════════════════════════════════════════════
    header_r = find_header_row(ws)
    if header_r is None:
        # Try looking with different keywords
        for r in range(1, min(10, max_row + 1)):
            row_text = set()
            for c in range(1, min(max_col + 1, 20)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    t = str(v).strip().lower()
                    row_text.add(t)
            if 'nombre' in row_text or any('nombre' in x for x in row_text):
                header_r = r
                break

    if header_r is None:
        # Last resort: check if R2 has header-like content
        for r in [2, 1, 3]:
            srow = sample.get(r, {})
            stext = ' '.join(srow.values()).lower()
            if 'nombre' in stext:
                header_r = r
                break

    if header_r is None:
        stats['errors'].append(f"Could not find headers in {filename} / {sheet_name}")
        return rows

    # Map columns
    col_map = {}
    for c in range(1, min(max_col + 1, 20)):
        v = ws.cell(row=header_r, column=c).value
        if v is None:
            continue
        vl = str(v).strip().lower()
        if vl in ('nombre', 'nombre ', 'nombres', 'nombres '):
            if 'nombre' not in col_map:
                col_map['nombre'] = c
        elif 'apellido' in vl:
            col_map['apellido'] = c
        elif 'cedula' in vl or 'cédula' in vl:
            col_map['cedula'] = c
        elif 'celular' in vl or 'telefono' in vl:
            col_map['celular'] = c
        elif 'llamada' in vl:
            col_map['llamada'] = c
        elif 'observaciones' in vl or 'observación' in vl or 'observacion' in vl:
            col_map['observaciones'] = c
        elif 'municipio' in vl:
            col_map['municipio'] = c
        elif 'parentesco' in vl:
            col_map['parentesco'] = c

    file_lider = extract_lider_from_filename(filename)

    # For simple municipality files, extract municipio from filename
    file_municipio = None
    if 'municipio' not in col_map:
        # If no municipio column, use filename
        fname_clean = os.path.splitext(filename)[0].rstrip('_').strip()
        # Only use as municipio if it looks like a municipality name (no person names, etc.)
        if file_lider is None and not any(kw in fname_clean.lower() for kw in
                                           ['planilla', 'consolidado', 'lista', 'referid', 'referent',
                                            'brigada', 'reunion', 'revisar', 'listado', 'campaña',
                                            'caminata', 'sin municipio']):
            file_municipio = fname_clean.upper()

    for r in range(header_r + 1, max_row + 1):
        nombre = clean_val(get_cell(ws, r, col_map.get('nombre', 99)))
        apellido = clean_val(get_cell(ws, r, col_map.get('apellido', 99)))
        cedula = clean_number(get_cell(ws, r, col_map.get('cedula', 99)))
        celular = clean_phone(get_cell(ws, r, col_map.get('celular', 99)))
        llamada = clean_val(get_cell(ws, r, col_map.get('llamada', 99)))
        obs = clean_val(get_cell(ws, r, col_map.get('observaciones', 99)))
        municipio = clean_val(get_cell(ws, r, col_map.get('municipio', 99)))
        parentesco = clean_val(get_cell(ws, r, col_map.get('parentesco', 99)))

        a_quien = build_full_name(nombre, apellido)

        if not a_quien and not celular:
            stats['rows_empty_skipped'] += 1
            continue

        final_obs = obs or llamada or 'SE ENVIO MENSAJE'
        final_municipio = municipio or file_municipio

        rows.append(make_row(
            nombre_lider=file_lider,
            a_quien=a_quien,
            num_doc=cedula,
            telefono=celular,
            relacion=parentesco,
            municipio=final_municipio,
            observaciones=final_obs,
        ))

    return rows


def is_data_sheet(sheet_name, wb):
    """Check if a sheet should be processed (skip lookup tables, etc.)."""
    sn = sheet_name.lower().strip()

    # Skip Hoja2 if it's a municipio/provincia lookup table
    if sn == 'hoja2':
        ws = wb[sheet_name]
        r1_c1 = ws.cell(row=1, column=1).value
        if r1_c1 and 'municipio' in str(r1_c1).lower():
            return False

    # Skip pivot table sheets
    if 'dinám' in sn or 'dinámica' in sn or 'dinßmica' in sn or 'pivot' in sn:
        return False

    # Skip "PLANILLA DE REGISTRO" (empty template)
    if 'planilla de registro' in sn:
        return False

    # Skip Hoja1 sheets that are just numbers (not data)
    if sn == 'hoja1':
        ws = wb[sheet_name]
        r1_c1 = ws.cell(row=1, column=1).value
        if r1_c1 is not None:
            s = str(r1_c1).strip()
            if s.replace('.', '').replace(',', '').isdigit():
                return False

    return True


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    all_rows = []

    # Collect all xlsx files recursively
    xlsx_files = []
    for root, dirs, files in os.walk(BASE_DIR):
        for f in files:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                xlsx_files.append(os.path.join(root, f))

    print(f"Found {len(xlsx_files)} Excel files to process\n")

    for filepath in sorted(xlsx_files):
        filename = os.path.basename(filepath)
        rel_path = os.path.relpath(filepath, BASE_DIR)

        # Skip our inspection files
        if filename in ('inspect_files.py', 'inspect_output.txt', 'inspect_output2.txt'):
            continue

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            stats['errors'].append(f"Cannot open {rel_path}: {e}")
            stats['files_skipped'] += 1
            continue

        file_rows = 0
        for sheet_name in wb.sheetnames:
            if not is_data_sheet(sheet_name, wb):
                stats['sheets_skipped'] += 1
                continue

            ws = wb[sheet_name]
            try:
                sheet_rows = detect_and_parse_sheet(ws, sheet_name, filename, filepath)
                all_rows.extend(sheet_rows)
                file_rows += len(sheet_rows)
                stats['sheets_processed'] += 1
            except Exception as e:
                stats['errors'].append(f"Error in {rel_path}/{sheet_name}: {e}")

        if file_rows > 0:
            print(f"  [OK] {rel_path}: {file_rows} rows")
            stats['files_processed'] += 1
        else:
            print(f"  [--] {rel_path}: 0 rows (skipped)")
            stats['files_skipped'] += 1

    # Write CSV
    print(f"\n{'='*60}")
    print(f"Writing {len(all_rows)} rows to {OUTPUT_CSV}")

    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)

    stats['rows_added'] = len(all_rows)

    # Print stats
    print(f"\n{'='*60}")
    print("ESTADÍSTICAS:")
    print(f"  Archivos procesados: {stats['files_processed']}")
    print(f"  Archivos sin datos:  {stats['files_skipped']}")
    print(f"  Hojas procesadas:    {stats['sheets_processed']}")
    print(f"  Hojas omitidas:      {stats['sheets_skipped']}")
    print(f"  Filas agregadas:     {stats['rows_added']}")
    print(f"  Filas vacías omitidas: {stats['rows_empty_skipped']}")

    if stats['errors']:
        print(f"\n  ERRORES ({len(stats['errors'])}):")
        for e in stats['errors']:
            print(f"    - {e}")

    print(f"\nCSV guardado en: {OUTPUT_CSV}")


if __name__ == '__main__':
    main()
