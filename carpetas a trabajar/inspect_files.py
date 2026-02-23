import openpyxl
import os

base_dir = r'c:\Users\Usuario\OneDrive\Escritorio\Pagina web\carpetas a trabajar'

# Investigate planilla files more deeply - need to find where the actual data rows start
planilla_files = [
    'Copia de planilla A.P cota NICOLÁS CALDERON.xlsx',
    'Cristian muñoz- planilla A.P cota.xlsx',
    'LISTADO CAMPAÑA CAMARA ALEX JUAN BERNAL.xlsx',
    'PLANILLA BASE DATOS ANDREA RODRIGUEZ RUBEN CACERES.xlsx',
    'Copia de planilla ALEJANDRO RAMIREZ.xlsx',
    'Consolidado Cota.xlsx',
    'Consolidado Cota No. 2.xlsx',
    'REUNION 20 ENERO COTA.xlsx',
    'BRIGADA SALUD ORAL 18 ENERO COTA.xlsx',
    'Brigada animal 25 de enero.xlsx',
    'REVISAR-BASE DE DATOS REFERIDOS DEPURADA.xlsx',
    'REVISAR-BASE.xlsx',
    'REVISAR-CAMINATAS SEÑORA DEISY .xlsx',
    'REVISAR-LISTADO CAMPAÑA CAMARA ALEX090226.xlsx',
    'Lista Ingrid Gonzalez.xlsx',
]

for f in planilla_files:
    path = os.path.join(base_dir, f)
    if not os.path.exists(path):
        print(f"NOT FOUND: {f}")
        continue
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            if ws.max_row < 2:
                continue
            print(f"\n{'='*70}")
            print(f"FILE: {f} | SHEET: {sname}")
            print(f"Rows: {ws.max_row} | Cols: {ws.max_column}")
            # Print first 10 rows to find where data starts
            for r in range(1, min(12, ws.max_row + 1)):
                row_vals = []
                for c in range(1, min(ws.max_column + 1, 20)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        row_vals.append(f"C{c}={str(v)[:35]}")
                if row_vals:
                    print(f"  R{r}: {row_vals}")
                else:
                    print(f"  R{r}: [empty]")
    except Exception as e:
        print(f"ERROR reading {f}: {e}")
